#!/usr/bin/env python3
"""
Complete Excel to JSON converter for FORTRESS Framework
Extracts all data from Fortress_Framework_v9.xlsx
"""

import json
import sys

def excel_to_json(excel_file, json_output):
    """Convert Excel file to JSON with all fields"""
    
    # Import openpyxl, install if needed
    try:
        import openpyxl
    except ImportError:
        print("📦 Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])
        import openpyxl
    
    print(f"📖 Reading {excel_file}...")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_file)
    
    # Use the "Fortress Framework" sheet
    if "Fortress Framework" in wb.sheetnames:
        ws = wb["Fortress Framework"]
    else:
        ws = wb.active
    
    print(f"   Sheet: {ws.title}")
    print(f"   Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    # Get headers from row 2 (row 1 might be title)
    headers = []
    header_row = 2
    
    # Find the header row
    for row in range(1, 5):
        first_cell = ws.cell(row=row, column=1).value
        if first_cell and 'Item Number' in str(first_cell):
            header_row = row
            break
    
    print(f"   Header row: {header_row}")
    
    # Extract headers
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col).value
        if header:
            headers.append(str(header).strip())
        else:
            headers.append(f"Column_{col}")
    
    print(f"   Headers: {headers[:5]}...")
    
    # Extract data
    items = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        item_number = ws.cell(row=row_idx, column=1).value
        
        if not item_number:
            continue
        
        item_number = str(item_number).strip()
        
        # Skip section headers (no dots or only one dot like "1" or "1.1")
        if '.' not in item_number or item_number.count('.') < 2:
            continue
        
        # Extract all columns
        for col_idx, header in enumerate(headers, start=1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            
            # Clean up the value
            if cell_value is None:
                row_data[header] = ""
            else:
                row_data[header] = str(cell_value).strip()
        
        # Parse mapped standards into array
        standards_str = row_data.get('Mapped Standards', '')
        standards_list = []
        if standards_str:
            # Split by comma and clean up
            standards_list = [s.strip() for s in standards_str.split(',') if s.strip()]
        
        # Create structured item
        item = {
            "item_number": row_data.get('Item Number', ''),
            "item_description": row_data.get('Item Description', ''),
            "super_section": row_data.get('Super Section', ''),
            "parent_section": row_data.get('Parent Section', ''),
            "tactic": row_data.get('Tactic (Goal)', ''),
            "technique": row_data.get("Technique (How it's done)", ''),
            "procedure": row_data.get('Procedure (Example)', ''),
            "compliance_rationale": row_data.get('Why (Compliance Rationale)', ''),
            "test_method": row_data.get('How (Test Method)', ''),
            "compliance_mappings": standards_list,
            "finding": row_data.get('Finding', ''),
            "recommendation": row_data.get('Recommendation', ''),
            "status": "active",
            "tags": [],
            "references": []
        }
        
        items.append(item)
    
    print(f"   Extracted {len(items)} items")
    
    # Write to JSON
    print(f"💾 Writing to {json_output}...")
    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump(items, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Success! Created {json_output}")
    print(f"   Total items: {len(items)}")
    
    # Show sample
    if items:
        print(f"\n📋 Sample item (first):")
        print(f"   Item Number: {items[0]['item_number']}")
        print(f"   Item Description: {items[0]['item_description']}")
        print(f"   Tactic: {items[0]['tactic']}")
        print(f"   Technique: {items[0]['technique']}")
        
        # Show sample from middle
        if len(items) > 100:
            sample = items[100]
            print(f"\n📋 Sample item (1.3.1 area):")
            print(f"   Item Number: {sample['item_number']}")
            print(f"   Item Description: {sample['item_description']}")
            print(f"   Finding: {sample['finding'][:80]}..." if len(sample['finding']) > 80 else f"   Finding: {sample['finding']}")
    
    return len(items)

def main():
    excel_file = "Fortress_Framework_v9.xlsx"
    json_output = "fortressframework.json"
    
    print("🏰 FORTRESS Framework Excel to JSON Converter")
    print("=" * 60)
    
    try:
        total = excel_to_json(excel_file, json_output)
        print("=" * 60)
        print(f"✅ Conversion complete! {total} items exported.")
        print(f"\n📁 Output file: {json_output}")
        print(f"🌐 Open navigator.html in a browser to view the data")
        
    except FileNotFoundError:
        print(f"❌ Error: {excel_file} not found")
        print("   Make sure the Excel file is in the current directory")
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

